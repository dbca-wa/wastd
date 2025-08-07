from django.shortcuts import render, redirect, get_object_or_404
from django.forms import inlineformset_factory
from .models import Incident, Uploaded_file, Species
from .forms import IncidentForm, UploadedFileForm
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import csv
import xlwt
from openpyxl import Workbook
from django.http import HttpResponse
from django.http import JsonResponse
from django.views.decorators.http import require_GET
from .decorators import superuser_or_data_curator_required
from django.contrib.auth.decorators import user_passes_test
from django.core.exceptions import PermissionDenied
from django.core.exceptions import ValidationError
from openpyxl import load_workbook
from datetime import datetime
from datetime import time
from openpyxl.utils.datetime import from_excel
from django.contrib.gis.geos import Point
from decimal import Decimal

INCIDENT_TYPE_MAP = {
    'stranding': 'Stranding',
    'entanglement': 'Entanglement',
    'entrapment': 'Entrapment',
    'vessel collision': 'Vessel collision',
    'unusual mortality event': 'Unusual mortality event',
    'hauled-out': 'Hauled-out'
}

SEX_MAP = {
    'F': 'Female',
    'M': 'Male',
    'U': 'Unknown',
    'f': 'Female',
    'm': 'Male',
    'u': 'Unknown'
}

CONDITION_MAP = {
    'Stage 1= alive': 'Stage 1 = alive',
    'Stage 2= fresh dead': 'Stage 2 = fresh dead',
    'Stage 3= mild decomposition': 'Stage 3 = mild decomposition',
    'Stage 4= advanced decomposition': 'Stage 4 = advanced decomposition',
    'Stage 5= mummified/skeletal': 'Stage 5 = mummified/skeletal',
    'unknown': 'Unknown'
}

OUTCOME_MAP = {
    'Dead': 'Died',
    'Euthanased': 'Euthanased',
    'Restranded and euthanased': 'Restranded and euthanased',
    'Refloated, fate unknown': 'Refloated, fate unknown',
    'Unknown': 'Unknown'
}

def round_decimal(value, places=2):
    """Round decimal to 2 decimal places"""
    if value is None:
        return None
    try:
        return Decimal(str(value)).quantize(Decimal('0.01'))
    except:
        return None

def user_in_marine_animal_incidents_group(user):
    return user.is_superuser or user.groups.filter(name='MARINE_ANIMAL_INCIDENTS') or user.groups.filter(name='data curator').exists()

@user_passes_test(user_in_marine_animal_incidents_group, login_url=None, redirect_field_name=None)
def incident_form(request, pk=None):
    """Handle incident creation and update"""
    if not user_in_marine_animal_incidents_group(request.user):
        raise PermissionDenied("You don't have permission to access this page")
    
    # Get incident instance if updating
    incident = get_object_or_404(Incident, pk=pk) if pk else None
    
    # Create file upload formset
    UploadedFileFormSet = inlineformset_factory(
        Incident, 
        Uploaded_file,
        form=UploadedFileForm,
        extra=0,
        can_delete=True,
    )

    if request.method == 'POST':
        form = IncidentForm(request.POST, instance=incident)
        formset = UploadedFileFormSet(request.POST, request.FILES, instance=incident)

        if form.is_valid() and formset.is_valid():
            try:
                # 1. Save the incident first
                incident = form.save()

                # 2. Save the file uploads
                formset.instance = incident
                formset.save()

                # 3. Return success response
                response_data = {
                    'status': 'success',
                    'message': 'Incident saved successfully'
                }
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse(response_data)
                else:
                    messages.success(request, 'Incident saved successfully')
                    return redirect('marine_mammal_incidents:incident_list')

            except Exception as e:
                error_data = {
                    'status': 'error',
                    'message': f'Error saving incident: {str(e)}',
                }
                return JsonResponse(error_data, status=400)
        else:
            # Handle form validation errors
            errors = {
                'status': 'error',
                'errors': {
                    'form_errors': form.errors,
                    'formset_errors': [f.errors for f in formset]
                }
            }
            return JsonResponse(errors, status=400)

    else:
        # GET request - display form
        form = IncidentForm(instance=incident)
        formset = UploadedFileFormSet(instance=incident)

    context = {
        'form': form,
        'formset': formset,
        'form_title': 'Update Incident' if pk else 'Create New Incident',
        'submit_button_text': 'Update' if pk else 'Create'
    }

    return render(request, 'marine_mammal_incidents/incident_form.html', context)


def incident_list(request):
    incidents = Incident.objects.all().order_by('-incident_date')
    
    paginator = Paginator(incidents, 30) 
    page = request.GET.get('page')
    
    try:
        incidents = paginator.page(page)
    except PageNotAnInteger:
        incidents = paginator.page(1)
    except EmptyPage:
        incidents = paginator.page(paginator.num_pages)
    
    context = {
        'incidents': incidents,
        'object_count': paginator.count,
        'is_paginated': incidents.has_other_pages(),
    }
    return render(request, 'marine_mammal_incidents/incident_list.html', context)

@superuser_or_data_curator_required
def export_form(request):
    try:
        species_list = Species.objects.all()
    except Exception as e:
        species_list = []
    return render(request, 'marine_mammal_incidents/export_form.html', {'species_list': species_list})

@superuser_or_data_curator_required
def export_data(request):

    incident_date_from = request.GET.get('incident_date_from')
    incident_date_to = request.GET.get('incident_date_to')
    species_id = request.GET.get('species')
    location_name = request.GET.get('location_name')
    file_format = request.GET.get('format', 'csv')

    incidents = Incident.objects.all()
    if incident_date_from:
        incidents = incidents.filter(incident_date__gte=incident_date_from)
    if incident_date_to:
        incidents = incidents.filter(incident_date__lte=incident_date_to)
    if species_id:
        incidents = incidents.filter(species_id=species_id)
    if location_name:
        incidents = incidents.filter(location_name__icontains=location_name)

    # Prepare data
    data = [['Species', 'Latitude', 'Longitude', 'ID', 'Incident Date', 'Incident Time', 
            'Species Confirmed Genetically', 'Location Name', 'Geo Location', 
            'Number of Animals', 'Mass Incident', 'Incident Type', 'Sex', 'Age Class', 
            'Length', 'Weight', 'Weight Is Estimated', 'Carcass Location Fate', 
            'Entanglement Gear', 'DBCA Staff Attended', 'Condition When Found', 
            'Outcome', 'Cause of Death', 'Photos Taken', 'Samples Taken', 
            'Post Mortem', 'Comments']]

    for incident in incidents:
        data.append([
            str(incident.species) if incident.species is not None else '',
            incident.geo_location.y if incident.geo_location else '',
            incident.geo_location.x if incident.geo_location else '',
            incident.id if incident.id is not None else '',
            incident.incident_date if incident.incident_date is not None else '',
            incident.incident_time if incident.incident_time is not None else '',
            incident.species_confirmed_genetically if incident.species_confirmed_genetically is not None else '',
            incident.location_name if incident.location_name is not None else '',
            incident.geo_location if incident.geo_location is not None else '',
            incident.number_of_animals if incident.number_of_animals is not None else '',
            incident.mass_incident if incident.mass_incident is not None else '',
            incident.incident_type if incident.incident_type is not None else '',
            incident.sex if incident.sex is not None else '',
            incident.age_class if incident.age_class is not None else '',
            incident.length if incident.length is not None else '',
            incident.weight if incident.weight is not None else '',
            incident.weight_is_estimated if incident.weight_is_estimated is not None else '',
            incident.carcass_location_fate if incident.carcass_location_fate is not None else '',
            incident.entanglement_gear if incident.entanglement_gear is not None else '',
            incident.DBCA_staff_attended if incident.DBCA_staff_attended is not None else '',
            incident.condition_when_found if incident.condition_when_found is not None else '',
            incident.outcome if incident.outcome is not None else '',
            incident.cause_of_death if incident.cause_of_death is not None else '',
            incident.photos_taken if incident.photos_taken is not None else '',
            incident.samples_taken if incident.samples_taken is not None else '',
            incident.post_mortem if incident.post_mortem is not None else '',
            incident.comments if incident.comments is not None else ''
        ])

    if file_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="incidents.csv"'
        writer = csv.writer(response)
        writer.writerows(data)
    elif file_format == 'xls':
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="incidents.xls"'
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Incidents')
        for row_num, row_data in enumerate(data):
            for col_num, cell_value in enumerate(row_data):
                ws.write(row_num, col_num, cell_value)
        wb.save(response)
    elif file_format == 'xlsx':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="incidents.xlsx"'
        wb = Workbook()
        ws = wb.active
        for row in data:
            ws.append(row)
        wb.save(response)
    else:
        return HttpResponse("Invalid file format")

    return response


@require_GET
def get_locations(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    locations = Incident.objects.filter(
        incident_date__range=[start_date, end_date]
    ).values_list('location_name', flat=True).distinct()
    
    return JsonResponse(list(locations), safe=False)

@superuser_or_data_curator_required
def import_incidents(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        try:
            excel_file = request.FILES['excel_file']
            wb = load_workbook(excel_file, data_only=True)
            ws = wb.active
            
            success_count = 0
            failed_rows = [] 
            error_messages = []
            
            # Save header row
            headers = [cell.value for cell in ws[1]]
            
            # Skip the header row
            for row in ws.iter_rows(min_row=2):
                try:
                    # Check required fields
                    if not row[0].value:  # species
                        raise ValidationError(f"Row {row[0].row}: Species name cannot be empty")
                        
                    species = Species.objects.get(scientific_name=row[0].value)
                    
                    # Process time field
                    cell = row[5]
                    incident_time = None
                    
                    if cell.value:
                        if isinstance(cell.value, (datetime, time)):
                            incident_time = cell.value.time() if isinstance(cell.value, datetime) else cell.value
                        elif isinstance(cell.value, str) and cell.value != '00:00:00':
                            try:
                                incident_time = datetime.strptime(cell.value, '%H:%M:%S').time()
                            except ValueError:
                                pass
                        elif isinstance(cell.value, float):
                            try:
                                # Convert Excel time value
                                dt = from_excel(cell.value)
                                incident_time = dt.time()
                            except:
                                pass
                    
                    # Create Incident instance
                    incident_date = None
                    if row[4].value:
                        if isinstance(row[4].value, datetime):
                            incident_date = row[4].value.date()
                        elif isinstance(row[4].value, str):
                            try:
                                incident_date = datetime.strptime(row[4].value, '%Y-%m-%d').date()
                            except ValueError:
                                pass
                    
                    # Check for duplicate incidents based on ID first
                    incident_id = row[3].value
                    if incident_id:
                        # Check if incident with this ID already exists
                        if Incident.objects.filter(id=incident_id).exists():
                            row_data = [cell.value for cell in row]
                            failed_rows.append({
                                'row_number': row[0].row,
                                'data': row_data,
                                'error': 'Duplicate ID - incident with this ID already exists in database'
                            })
                            error_messages.append(f"Row {row[0].row}: duplicate ID - incident with this ID already exists in database")
                            continue
                    
                    # Check for duplicate incidents based on other fields
                    geo_location = Point(row[2].value, row[1].value) if row[1].value and row[2].value else None
                    incident_type = INCIDENT_TYPE_MAP.get(str(row[11].value).lower(), 'Stranding')
                    
                    # Check if this incident already exists based on other criteria
                    duplicate_query = Incident.objects.filter(
                        species=species,
                        incident_date=incident_date,
                        incident_time=incident_time,
                        incident_type=incident_type
                    )
                    
                    # If geo_location exists, add it to the duplicate check
                    if geo_location:
                        duplicate_query = duplicate_query.filter(geo_location=geo_location)
                    
                    if duplicate_query.exists():
                        # This is a duplicate, skip it
                        row_data = [cell.value for cell in row]
                        failed_rows.append({
                            'row_number': row[0].row,
                            'data': row_data,
                            'error': 'Duplicate incident - already exists in database'
                        })
                        error_messages.append(f"Row {row[0].row}: duplicate incident - already exists in database")
                        continue
                    
                    incident = Incident(
                        species=species,
                        geo_location=geo_location,
                        incident_date=incident_date,
                        incident_time=incident_time,
                        species_confirmed_genetically=row[6].value == 'Y' if row[6].value else False,
                        location_name=row[7].value if row[7].value else '',
                        number_of_animals=int(row[9].value) if row[9].value else 1,
                        mass_incident=row[10].value == 'Y' if row[10].value else False,
                        incident_type=incident_type,
                        sex=SEX_MAP.get(str(row[12].value).strip(), 'Unknown'),
                        age_class=row[13].value if row[13].value else 'Unknown',
                        length=round_decimal(row[14].value),
                        weight=round_decimal(row[15].value),
                        weight_is_estimated=row[16].value == 'Y' if row[16].value else False,
                        carcass_location_fate=row[17].value if row[17].value else '',
                        entanglement_gear=row[18].value if row[18].value else '',
                        DBCA_staff_attended=row[19].value == 'Y' if row[19].value else False,
                        condition_when_found=CONDITION_MAP.get(str(row[20].value).strip(), 'Unknown'),
                        outcome=OUTCOME_MAP.get(str(row[21].value).strip(), 'Unknown'),
                        cause_of_death=row[22].value if row[22].value else '',
                        photos_taken=row[23].value == 'Y' if row[23].value else False,
                        samples_taken=row[24].value == 'Y' if row[24].value else False,
                        post_mortem=row[25].value == 'Y' if row[25].value else False,
                        comments=row[26].value if row[26].value else ''
                    )
                    
                    incident.full_clean()
                    incident.save()
                    success_count += 1
                    
                except (ValidationError, Exception) as e:
                    # Save failed row data and error message
                    row_data = [cell.value for cell in row]
                    failed_rows.append({
                        'row_number': row[0].row,
                        'data': row_data,
                        'error': str(e)
                    })
                    error_messages.append(f"行 {row[0].row}: {str(e)}")
                    continue
            
            # If there are failed rows, create an error report Excel file
            if failed_rows:
                wb_error = Workbook()
                ws_error = wb_error.active
                ws_error.title = "Failed Records"
                
                # Write header row
                headers.append("Error Message")  # Add error message column
                ws_error.append(headers)
                
                # Write failed row data
                for failed_row in failed_rows:
                    row_data = failed_row['data']
                    row_data.append(failed_row['error'])  # Add error message
                    ws_error.append(row_data)
                
                # Generate file name
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f'failed_imports_{timestamp}.xlsx'
                
                # 保存并返回错误报告文件
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                wb_error.save(response)
                
                messages.warning(
                    request, 
                    f'Successfully imported {success_count} records.'
                    f'Failed {len(failed_rows)} records, please download the error report for correction.'
                )
                return response
            
            messages.success(request, f'Successfully imported {success_count} records.')
            
        except Exception as e:
            messages.error(request, f'Import failed: {str(e)}')
            
    return render(request, 'marine_mammal_incidents/import_form.html')

